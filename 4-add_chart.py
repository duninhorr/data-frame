from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference




#1 lê pasta de trablho e planilha 

Wb =load_workbook("data/pivot_table.xlsx")

sheet = Wb["relatorio"]

#2 refering to the lines and columns


min_column = Wb.active.min_column
max_column = Wb.active.max_column
min_row = Wb.active.min_row
max_row = Wb.active.max_row


# 3 adicons for data and categories in graphics

barChart = BarChart()


data= Reference(
    sheet,
    min_col=min_column + 1,
    max_col=max_column,
    min_row=min_row , 
    max_row=max_row
)

caregories= Reference(
    sheet,
    min_col=min_column,
    max_col=max_column,
    min_row=min_row + 1, 
    max_row=max_row
)

barChart.add_data(data,titles_from_data = True)
barChart.set_categories(caregories)

#4 criando  o graphics

sheet.add_chart(barChart,"B10")
barChart.title = "vendas por fabricante"
barChart.style = 3

# 5 salvando meus workbook 

Wb.save("data/barchart.xlsx")
 
 